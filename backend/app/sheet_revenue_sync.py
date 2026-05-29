"""Google Sheets(정산 관리) ↔ boda.db 수납·정산 맞추기.

- 수업료 관리 → billing_unit=monthly
- 회당 수금표 → billing_unit=per_session (이번달 횟수·금액)
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .models import LessonEnrollment, MonthlyPaymentRecord, StudentProfile, TeacherProfile, User
from .payment_pricing import recompute_payment_final_amount
from .settlement_sync import sync_settlements_from_payments

RECURRING_CHARGE_LABELS = frozenset({"정기결제", "첫달결제"})
MONTH_TITLE_PATTERN = re.compile(r"(\d{4})년\s*(\d{1,2})월")
SHEET_MEMO_PREFIX = "[sheet]"

TUITION_COL_TEACHER = 2
TUITION_COL_STUDENT = 15
TUITION_COL_REGULAR_TUITION = 33
TUITION_COL_FIRST_PAYMENT = 37
TUITION_COL_LESSON_START = 31
TUITION_COL_LESSON_END = 40

SESSION_COL_TEACHER = 2
SESSION_COL_STUDENT = 3
SESSION_COL_TRIAL_DATE = 12
SESSION_COL_LESSON_START = 13
SESSION_COL_LESSON_END = 14


@dataclass(frozen=True)
class SheetPaymentRow:
    billing_month: str
    teacher_name: str
    student_name: str
    billing_unit: str
    amount: int
    total_sessions: int
    payment_method: Optional[str] = None
    charge_label: Optional[str] = None


def default_workbook_path() -> Path:
    root = Path(__file__).resolve().parents[2]
    for candidate in (root / "sheet.xlsx", root.parent / "sheet.xlsx"):
        if candidate.is_file():
            return candidate
    return root / "sheet.xlsx"


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _int(value: Any) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _month_from_value(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        dt = value.date() if isinstance(value, datetime) else value
        return f"{dt.year:04d}-{dt.month:02d}"
    text = _text(value)
    match = MONTH_TITLE_PATTERN.search(text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    return ""


def _normalize_person_name(value: str) -> str:
    text = _text(value)
    text = re.sub(r"\([^)]*\)", "", text)
    return text.replace(" ", "").lower()


def _normalize_payment_method(value: str | None) -> str | None:
    payment_method = _text(value)
    if not payment_method or payment_method in {"ex)"}:
        return None
    if payment_method.lower() == "cms":
        return "CMS"
    if payment_method in {"O", "X"}:
        return "결제"
    return payment_method


def _parse_month_bounds(month_key: str) -> tuple[date, date] | None:
    try:
        year_text, month_text = month_key.split("-")
        year, month = int(year_text), int(month_text)
    except ValueError:
        return None
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1
    return date(year, month, 1), date(next_year, next_month, 1) - timedelta(days=1)


def _is_lesson_active_in_month(lesson_start: date | None, lesson_end: date | None, month_key: str) -> bool:
    bounds = _parse_month_bounds(month_key)
    if not bounds:
        return False
    month_start, month_end = bounds
    if lesson_start and lesson_start > month_end:
        return False
    if lesson_end and lesson_end < month_start:
        return False
    return True


def _header_month_column_map(header_row: tuple) -> dict[int, str]:
    month_columns: dict[int, str] = {}
    for index, value in enumerate(header_row):
        if isinstance(value, (date, datetime)):
            month_columns[index] = _month_from_value(value)
    return month_columns


def _month_charge_amount(row: tuple, column_index: int, regular_tuition: float) -> tuple[float, str | None]:
    cell = row[column_index] if len(row) > column_index else None
    if isinstance(cell, (int, float)) and cell > 0:
        return float(cell), "amount"

    label = _text(cell)
    if label in RECURRING_CHARGE_LABELS and regular_tuition > 0:
        if label == "첫달결제":
            first_amount = row[TUITION_COL_FIRST_PAYMENT] if len(row) > TUITION_COL_FIRST_PAYMENT else None
            if isinstance(first_amount, (int, float)) and first_amount > 0:
                return float(first_amount), label
        return regular_tuition, label
    return 0.0, label or None


def parse_tuition_payments(worksheet, *, billing_month: Optional[str] = None) -> list[SheetPaymentRow]:
    records: list[SheetPaymentRow] = []
    header_row = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
    month_columns = _header_month_column_map(header_row)

    for row in worksheet.iter_rows(min_row=5, values_only=True):
        payment_method = _normalize_payment_method(row[1] if len(row) > 1 else None)
        teacher_name = _text(row[TUITION_COL_TEACHER] if len(row) > TUITION_COL_TEACHER else None)
        student_name = _text(row[TUITION_COL_STUDENT] if len(row) > TUITION_COL_STUDENT else None)
        if not payment_method or not teacher_name or not student_name or student_name.startswith("▶"):
            continue

        regular_tuition = _float(row[TUITION_COL_REGULAR_TUITION] if len(row) > TUITION_COL_REGULAR_TUITION else None)
        lesson_start = _date(row[TUITION_COL_LESSON_START] if len(row) > TUITION_COL_LESSON_START else None)
        lesson_end = _date(row[TUITION_COL_LESSON_END] if len(row) > TUITION_COL_LESSON_END else None)

        for column_index, month_key in month_columns.items():
            if billing_month and month_key != billing_month:
                continue
            if not month_key or not _is_lesson_active_in_month(lesson_start, lesson_end, month_key):
                continue

            amount, charge_label = _month_charge_amount(row, column_index, regular_tuition)
            if amount <= 0:
                continue

            records.append(
                SheetPaymentRow(
                    billing_month=month_key,
                    teacher_name=teacher_name,
                    student_name=student_name,
                    billing_unit="monthly",
                    amount=int(round(amount)),
                    total_sessions=0,
                    payment_method=payment_method,
                    charge_label=charge_label,
                )
            )
    return records


def parse_session_payments(worksheet, *, billing_month: Optional[str] = None) -> list[SheetPaymentRow]:
    records: list[SheetPaymentRow] = []
    header_row = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
    month_columns = _header_month_column_map(header_row)
    display_month = _month_from_value(worksheet["K2"].value)

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        teacher_name = _text(row[SESSION_COL_TEACHER] if len(row) > SESSION_COL_TEACHER else None)
        student_name = _text(row[SESSION_COL_STUDENT] if len(row) > SESSION_COL_STUDENT else None)
        if not teacher_name or not student_name:
            continue

        display_amount = _float(row[11] if len(row) > 11 else None)
        display_sessions = _int(row[10] if len(row) > 10 else None)
        first_payment_amount = _float(row[15] if len(row) > 15 else None)
        payment_method = _normalize_payment_method(row[1] if len(row) > 1 else None)
        lesson_start = _date(row[SESSION_COL_LESSON_START] if len(row) > SESSION_COL_LESSON_START else None)
        lesson_end = _date(row[SESSION_COL_LESSON_END] if len(row) > SESSION_COL_LESSON_END else None)

        month_activity: dict[str, dict[str, Any]] = {}

        def ensure_month(month_key: str) -> dict[str, Any]:
            return month_activity.setdefault(month_key, {"amount": 0.0, "sessions": 0, "label": None})

        for column_index, month_key in month_columns.items():
            if display_month and month_key > display_month:
                continue
            if billing_month and month_key != billing_month:
                continue

            cell = row[column_index] if len(row) > column_index else None
            if cell in (None, ""):
                continue

            entry = ensure_month(month_key)
            if isinstance(cell, (date, datetime)):
                continue

            label = _text(cell)
            if label in RECURRING_CHARGE_LABELS:
                if label == "첫달결제" and first_payment_amount > 0:
                    amount = first_payment_amount
                elif month_key == display_month and display_amount > 0:
                    amount = display_amount
                else:
                    amount = 0.0
                if amount > 0:
                    entry["amount"] = max(entry["amount"], amount)
                    entry["sessions"] = max(entry["sessions"], display_sessions)
                    entry["label"] = label
            elif isinstance(cell, (int, float)) and cell > 0:
                entry["amount"] = max(entry["amount"], float(cell))
                entry["sessions"] = max(entry["sessions"], display_sessions)
                entry["label"] = "amount"

        # K/L열(이번달 횟수·금액) = 조회 기준월(K2) 수납의 정본
        if display_month and (display_amount > 0 or display_sessions > 0):
            entry = ensure_month(display_month)
            if display_amount > 0:
                entry["amount"] = display_amount
            if display_sessions > 0:
                entry["sessions"] = display_sessions

        for month_key, entry in month_activity.items():
            if billing_month and month_key != billing_month:
                continue
            if entry["amount"] <= 0 and entry["sessions"] <= 0:
                continue
            if not _is_lesson_active_in_month(lesson_start, lesson_end, month_key):
                continue

            records.append(
                SheetPaymentRow(
                    billing_month=month_key,
                    teacher_name=teacher_name,
                    student_name=student_name,
                    billing_unit="per_session",
                    amount=int(round(entry["amount"])),
                    total_sessions=max(0, _int(entry["sessions"])),
                    payment_method=payment_method,
                    charge_label=entry.get("label"),
                )
            )
    return records


def load_sheet_payments(
    workbook_path: Path | str,
    *,
    billing_month: Optional[str] = None,
) -> list[SheetPaymentRow]:
    workbook = load_workbook(filename=str(workbook_path), data_only=True, read_only=True)
    try:
        tuition = parse_tuition_payments(workbook["수업료 관리"], billing_month=billing_month)
        session = parse_session_payments(workbook["회당 수금표"], billing_month=billing_month)
    finally:
        workbook.close()

    session_keys = {
        (row.billing_month, _normalize_person_name(row.teacher_name), _normalize_person_name(row.student_name))
        for row in session
    }
    monthly = [
        row
        for row in tuition
        if (row.billing_month, _normalize_person_name(row.teacher_name), _normalize_person_name(row.student_name))
        not in session_keys
    ]
    return monthly + session


def _teacher_name_map(db: Session) -> dict[str, tuple[int, str]]:
    mapping: dict[str, tuple[int, str]] = {}
    rows = (
        db.query(TeacherProfile.id, User.name)
        .join(User, User.id == TeacherProfile.user_id)
        .all()
    )
    for teacher_id, name in rows:
        key = _normalize_person_name(name or "")
        if key:
            mapping[key] = (int(teacher_id), name or "")
    return mapping


def _student_name_map(db: Session) -> dict[str, tuple[int, str]]:
    mapping: dict[str, tuple[int, str]] = {}
    rows = (
        db.query(StudentProfile.id, User.name)
        .join(User, User.id == StudentProfile.user_id)
        .all()
    )
    for student_id, name in rows:
        key = _normalize_person_name(name or "")
        if key:
            mapping[key] = (int(student_id), name or "")
    return mapping


def _resolve_teacher_id(name_map: dict[str, tuple[int, str]], sheet_name: str) -> Optional[int]:
    key = _normalize_person_name(sheet_name)
    if key in name_map:
        return name_map[key][0]
    for norm, (teacher_id, _) in name_map.items():
        if key.startswith(norm) or norm.startswith(key):
            return teacher_id
    return None


def _resolve_student_id(name_map: dict[str, tuple[int, str]], sheet_name: str) -> Optional[int]:
    key = _normalize_person_name(sheet_name)
    if key in name_map:
        return name_map[key][0]
    for norm, (student_id, _) in name_map.items():
        if key in norm or norm in key:
            return student_id
    return None


def _list_enrollments(db: Session, *, teacher_id: int, student_id: int) -> list[LessonEnrollment]:
    return (
        db.query(LessonEnrollment)
        .filter(LessonEnrollment.teacher_id == teacher_id, LessonEnrollment.student_id == student_id)
        .order_by(LessonEnrollment.start_date.desc().nullslast(), LessonEnrollment.id.desc())
        .all()
    )


def _pick_enrollment(
    enrollments: list[LessonEnrollment],
    *,
    used_ids: set[int],
) -> Optional[LessonEnrollment]:
    open_rows = [row for row in enrollments if row.id not in used_ids]
    if not open_rows:
        return None

    active = [row for row in open_rows if not row.end_date and not row.cancelled_at]
    if active:
        return active[0]
    return open_rows[0]


def _payment_tag_from_charge(charge_label: Optional[str]) -> str:
    if charge_label == "첫달결제":
        return "first_month"
    if charge_label in RECURRING_CHARGE_LABELS or charge_label == "amount":
        return "regular"
    return "regular"


def _apply_sheet_row_to_payment(
    db: Session,
    enrollment: LessonEnrollment,
    sheet_row: SheetPaymentRow,
) -> tuple[bool, str]:
    payment = (
        db.query(MonthlyPaymentRecord)
        .filter(
            MonthlyPaymentRecord.enrollment_id == enrollment.id,
            MonthlyPaymentRecord.billing_month == sheet_row.billing_month,
        )
        .first()
    )
    if not payment:
        payment = MonthlyPaymentRecord(
            billing_month=sheet_row.billing_month,
            enrollment_id=enrollment.id,
            student_id=enrollment.student_id,
            teacher_id=enrollment.teacher_id,
            billing_unit=sheet_row.billing_unit,
            total_sessions=0,
            completed_sessions=0,
            base_amount=0,
            special_amount=0,
            refund_amount=0,
            final_amount=0,
            commission_rate=float(enrollment.current_commission_rate or 60.0),
            payment_status="paid" if sheet_row.amount > 0 else "unpaid",
            payment_tag="unpaid",
            memo=None,
        )
        db.add(payment)
        db.flush()

    before = (
        payment.billing_unit,
        payment.total_sessions,
        payment.completed_sessions,
        payment.base_amount,
        payment.final_amount,
    )

    payment.billing_unit = sheet_row.billing_unit
    payment.commission_rate = float(enrollment.current_commission_rate or payment.commission_rate or 60.0)
    payment.memo = f"{SHEET_MEMO_PREFIX} {sheet_row.billing_unit}"

    if sheet_row.billing_unit == "per_session":
        sessions = max(0, sheet_row.total_sessions)
        payment.total_sessions = sessions
        payment.completed_sessions = sessions
        payment.base_amount = sheet_row.amount
        payment.special_amount = 0
        payment.refund_amount = 0
        recompute_payment_final_amount(payment, priced_amount=sheet_row.amount)
        enrollment.price_type = "session"
    else:
        payment.total_sessions = 0
        payment.completed_sessions = 0
        payment.base_amount = sheet_row.amount
        payment.special_amount = 0
        payment.refund_amount = 0
        recompute_payment_final_amount(payment, priced_amount=sheet_row.amount)
        if not enrollment.price_type:
            enrollment.price_type = "monthly"

    payment.payment_tag = _payment_tag_from_charge(sheet_row.charge_label)
    if sheet_row.amount > 0:
        payment.payment_status = "paid"
    elif payment.payment_status != "paid":
        payment.payment_status = "unpaid"

    if sheet_row.payment_method:
        enrollment.payment_method = sheet_row.payment_method

    after = (
        payment.billing_unit,
        payment.total_sessions,
        payment.completed_sessions,
        payment.base_amount,
        payment.final_amount,
    )
    changed = before != after
    return changed, "updated" if changed else "unchanged"


def sync_payment_records_from_workbook(
    db: Session,
    *,
    workbook_path: Path | str | None = None,
    billing_month: Optional[str] = None,
    min_month: Optional[str] = None,
) -> dict[str, Any]:
    path = Path(workbook_path) if workbook_path else default_workbook_path()
    if not path.is_file():
        return {"ok": False, "error": f"시트 파일을 찾을 수 없습니다: {path}", "updated": 0}

    sheet_rows = load_sheet_payments(path, billing_month=billing_month)
    if min_month:
        sheet_rows = [row for row in sheet_rows if row.billing_month >= min_month]
    teacher_map = _teacher_name_map(db)
    student_map = _student_name_map(db)

    updated = 0
    skipped: list[str] = []
    used_enrollment_ids: set[int] = set()
    months_touched: set[str] = set()

    grouped: dict[tuple[str, str, str, str], list[SheetPaymentRow]] = {}
    for row in sheet_rows:
        key = (
            row.billing_month,
            _normalize_person_name(row.teacher_name),
            _normalize_person_name(row.student_name),
            row.billing_unit,
        )
        grouped.setdefault(key, []).append(row)

    for (_month, _teacher_key, _student_key, unit), rows in grouped.items():
        rows = sorted(rows, key=lambda item: item.amount, reverse=True)
        for sheet_row in rows:
            months_touched.add(sheet_row.billing_month)
            teacher_id = _resolve_teacher_id(teacher_map, sheet_row.teacher_name)
            student_id = _resolve_student_id(student_map, sheet_row.student_name)
            if not teacher_id or not student_id:
                skipped.append(f"{sheet_row.teacher_name}/{sheet_row.student_name} (미매칭)")
                continue

            enrollments = _list_enrollments(db, teacher_id=teacher_id, student_id=student_id)
            if not enrollments:
                skipped.append(f"{sheet_row.teacher_name}/{sheet_row.student_name} (수업 없음)")
                continue

            enrollment = _pick_enrollment(enrollments, used_ids=used_enrollment_ids)
            if not enrollment:
                skipped.append(f"{sheet_row.teacher_name}/{sheet_row.student_name} (수업 매칭 실패)")
                continue

            used_enrollment_ids.add(int(enrollment.id))
            changed, _ = _apply_sheet_row_to_payment(db, enrollment, sheet_row)
            if changed:
                updated += 1

    settlement_updates = 0
    for month in sorted(months_touched):
        settlement_updates += sync_settlements_from_payments(db, billing_month=month)

    db.flush()
    return {
        "ok": True,
        "workbook": str(path),
        "sheet_rows": len(sheet_rows),
        "payment_rows_updated": updated,
        "settlement_rows_updated": settlement_updates,
        "months": sorted(months_touched),
        "skipped": skipped[:30],
        "skipped_count": len(skipped),
    }
