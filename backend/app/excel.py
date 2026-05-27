from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

SUPPORTED_HEADERS = {
    "date": "transaction_date",
    "날짜": "transaction_date",
    "category": "category",
    "분류": "category",
    "title": "title",
    "항목명": "title",
    "description": "title",
    "내용": "title",
    "amount": "amount",
    "금액": "amount",
    "payer": "payer",
    "결제자": "payer",
    "memo": "memo",
    "메모": "memo",
}

REQUIRED_COLUMNS = {"transaction_date", "category", "title", "amount", "payer"}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _parse_date(value: Any):
    if value is None or value == "":
        raise ValueError("날짜가 비어 있습니다.")
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    raise ValueError("날짜 형식이 올바르지 않습니다.")


def _parse_amount(value: Any) -> float:
    if value is None or value == "":
        raise ValueError("금액이 비어 있습니다.")
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError as exc:
        raise ValueError("금액은 숫자여야 합니다.") from exc


def parse_excel_rows(file_bytes: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], ["엑셀 파일에 데이터가 없습니다."]

    headers = rows[0]
    header_map: Dict[int, str] = {}
    for index, header in enumerate(headers):
        normalized = SUPPORTED_HEADERS.get(_normalize_header(header))
        if normalized:
            header_map[index] = normalized

    missing = REQUIRED_COLUMNS - set(header_map.values())
    if missing:
        return [], [f"필수 컬럼이 없습니다: {', '.join(sorted(missing))}"]

    parsed_items: List[Dict[str, Any]] = []
    skipped_rows: List[str] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue

        item: Dict[str, Any] = {"memo": ""}
        try:
            for index, field_name in header_map.items():
                item[field_name] = row[index] if index < len(row) else None

            parsed_item = {
                "title": str(item["title"]).strip(),
                "category": str(item["category"]).strip(),
                "amount": _parse_amount(item["amount"]),
                "transaction_date": _parse_date(item["transaction_date"]),
                "payer": str(item["payer"]).strip(),
                "memo": "" if item.get("memo") is None else str(item["memo"]).strip(),
            }

            if not parsed_item["title"] or not parsed_item["category"] or not parsed_item["payer"]:
                raise ValueError("필수 텍스트 값이 비어 있습니다.")

            parsed_items.append(parsed_item)
        except ValueError as exc:
            skipped_rows.append(f"{row_number}행: {exc}")

    return parsed_items, skipped_rows
