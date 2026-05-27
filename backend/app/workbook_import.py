from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .enrollment_sync import sync_teachers_and_enrollments
from .student_payment_sync import sync_student_payment_methods
from .models import (
    DataSync,
    MonthlySettlement,
    PaymentMethodRevenue,
    ProductPrice,
    RateTableRow,
    SessionCollection,
    SessionSettlement,
    StudentRecord,
    TeacherProfile,
    TeacherSettlement,
    TrialLesson,
    TuitionRecord,
    TuitionStudentMonth,
)

PAYMENT_METHOD_SKIP = {"ex)"}
PAYMENT_METHOD_ORDER = ("계좌이체", "카드", "납부자", "결제", "CMS")
RECURRING_CHARGE_LABELS = frozenset({"정기결제", "첫달결제"})

MONTH_TITLE_PATTERN = re.compile(r"(\d{4})년\s*(\d{1,2})월")

# 수업료 관리 시트 (0-based). AM열 = 시범수업 날짜
TUITION_COL_TEACHER = 2
TUITION_COL_STUDENT = 15
TUITION_COL_LESSON_START = 31
TUITION_COL_TRIAL_DATE = 38
TUITION_COL_LESSON_END = 40

# 회당 수금표 (0-based). M열 = 시범수업 날짜
SESSION_COL_TEACHER = 2
SESSION_COL_STUDENT = 3
SESSION_COL_TRIAL_DATE = 12
SESSION_COL_LESSON_START = 13
SESSION_COL_LESSON_END = 14

PRODUCT_PRICE_TABLES: tuple[tuple[str, tuple[tuple[str, int], ...]], ...] = (
    (
        "35% 할인",
        (
            ("고등 주1회 60분", 150000),
            ("고등 주1회 90분", 200000),
            ("고등 주1회 120분", 240000),
            ("고등 주2회 60분", 250000),
            ("고등 주2회 90분", 350000),
            ("고등 주2회 120분", 430000),
            ("고등 주3회 60분", 350000),
            ("고등 주3회 90분", 500000),
            ("고등 주3회 120분", 630000),
            ("중등 주1회 60분", 140000),
            ("중등 주1회 90분", 180000),
            ("중등 주1회 120분", 230000),
            ("중등 주2회 60분", 240000),
            ("중등 주2회 90분", 320000),
            ("중등 주2회 120분", 410000),
            ("중등 주3회 60분", 330000),
            ("중등 주3회 90분", 440000),
            ("중등 주3회 120분", 580000),
            ("초등 주1회 60분", 200000),
            ("초등 주2회 60분", 330000),
            ("초등 주3회 60분", 450000),
            ("고등 주2회 개별진도", 440000),
            ("고등 주3회 개별진도", 580000),
            ("고등 주4회 개별진도", 700000),
            ("중등 주2회 개별진도", 400000),
            ("중등 주3회 개별진도", 550000),
            ("중등 주4회 개별진도", 680000),
        ),
    ),
    (
        "17% 할인",
        (
            ("고등 주1회 60분", 210000),
            ("고등 주1회 90분", 250000),
            ("고등 주1회 120분", 300000),
            ("고등 주2회 60분", 400000),
            ("고등 주2회 90분", 440000),
            ("고등 주2회 120분", 540000),
            ("고등 주3회 60분", 550000),
            ("고등 주3회 90분", 630000),
            ("고등 주3회 120분", 790000),
            ("중등 주1회 60분", 190000),
            ("중등 주1회 90분", 230000),
            ("중등 주1회 120분", 290000),
            ("중등 주2회 60분", 320000),
            ("중등 주2회 90분", 400000),
            ("중등 주2회 120분", 520000),
            ("중등 주3회 60분", 440000),
            ("중등 주3회 90분", 550000),
            ("중등 주3회 120분", 730000),
            ("초등 주1회 60분", 170000),
            ("초등 주2회 60분", 280000),
            ("초등 주3회 60분", 380000),
            ("고등 주2회 개별진도", 440000),
            ("고등 주3회 개별진도", 580000),
            ("고등 주4회 개별진도", 700000),
            ("중등 주2회 개별진도", 400000),
            ("중등 주3회 개별진도", 550000),
            ("중등 주4회 개별진도", 680000),
        ),
    ),
    (
        "회당 단가표",
        (
            ("고등 주1회 60분", 55000),
            ("고등 주1회 90분", 65000),
            ("고등 주1회 120분", 75000),
            ("고등 주2회 60분", 50000),
            ("고등 주2회 90분", 55000),
            ("고등 주2회 120분", 70000),
            ("고등 주3회 60분", 50000),
            ("고등 주3회 90분", 55000),
            ("고등 주3회 120분", 70000),
            ("중등 주1회 60분", 48000),
            ("중등 주1회 90분", 58000),
            ("중등 주1회 120분", 70000),
            ("중등 주2회 60분", 40000),
            ("중등 주2회 90분", 50000),
            ("중등 주2회 120분", 65000),
            ("중등 주3회 60분", 37000),
            ("중등 주3회 90분", 46000),
            ("중등 주3회 120분", 61000),
            ("초등 주1회 60분", 42000),
            ("초등 주2회 60분", 35000),
            ("초등 주3회 60분", 32000),
        ),
    ),
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").replace("%", "").strip()
    if not cleaned:
        return 0.0
    return float(cleaned)


def _int(value: Any) -> int:
    return int(round(_float(value)))


def _date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        serial = float(value)
        if 30000 <= serial <= 60000:
            from openpyxl.utils.datetime import from_excel

            try:
                return from_excel(serial).date()
            except (ValueError, OverflowError):
                return None
        return None
    text = _text(value)
    normalized = re.sub(r"\s+", "", text)
    for candidate in (text, normalized):
        for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d.", "%Y%m%d"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if match:
        year, month, day = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
        try:
            return date(year, month, day)
        except ValueError:
            return None
    return None


def _month_from_value(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        dt = value.date() if isinstance(value, datetime) else value
        return f"{dt.year:04d}-{dt.month:02d}"
    text = _text(value)
    match = MONTH_TITLE_PATTERN.search(text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    return ""


def _normalize_payment_method(value: str | None) -> str | None:
    payment_method = _text(value)
    if not payment_method or payment_method in PAYMENT_METHOD_SKIP:
        return None
    if payment_method.lower() == "cms":
        return "CMS"
    if payment_method in {"O", "X"}:
        return "결제"
    return payment_method


def _parse_month_bounds(month_key: str) -> tuple[date, date] | None:
    parsed = _parse_month(month_key)
    if not parsed:
        return None
    year, month = parsed
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1
    month_start = date(year, month, 1)
    month_end = date(next_year, next_month, 1) - timedelta(days=1)
    return month_start, month_end


def _parse_month(month_key: str) -> tuple[int, int] | None:
    try:
        year_text, month_text = month_key.split("-")
        return int(year_text), int(month_text)
    except ValueError:
        return None


def _is_lesson_active_in_month(
    lesson_start: date | None,
    lesson_end: date | None,
    month_key: str,
) -> bool:
    bounds = _parse_month_bounds(month_key)
    if not bounds:
        return False
    month_start, month_end = bounds
    if lesson_start and lesson_start > month_end:
        return False
    if lesson_end and lesson_end < month_start:
        return False
    return True


def _month_charge_amount(row: tuple, column_index: int, regular_tuition: float) -> tuple[float, str | None]:
    cell = row[column_index] if len(row) > column_index else None
    if isinstance(cell, (int, float)) and cell > 0:
        return float(cell), "amount"

    label = _text(cell)
    if label in RECURRING_CHARGE_LABELS and regular_tuition > 0:
        if label == "첫달결제":
            first_amount = row[37] if len(row) > 37 else None
            if isinstance(first_amount, (int, float)) and first_amount > 0:
                return float(first_amount), label
        return regular_tuition, label

    return 0.0, label or None


def _tuition_student_months(worksheet) -> list[TuitionStudentMonth]:
    records: list[TuitionStudentMonth] = []
    header_row = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
    month_columns: dict[int, str] = {}
    for index, value in enumerate(header_row):
        if isinstance(value, (date, datetime)):
            month_columns[index] = _month_from_value(value)

    for row in worksheet.iter_rows(min_row=5, values_only=True):
        payment_method = _normalize_payment_method(row[1] if len(row) > 1 else None)
        student_name = _text(row[15] if len(row) > 15 else None)
        teacher_name = _text(row[2] if len(row) > 2 else None)
        if not payment_method or not student_name or student_name.startswith("▶") or not teacher_name:
            continue

        regular_value = row[33] if len(row) > 33 else None
        regular_tuition = float(regular_value) if isinstance(regular_value, (int, float)) else 0.0
        lesson_start = _date(row[31] if len(row) > 31 else None)
        lesson_end = _date(row[40] if len(row) > 40 else None)

        for column_index, month_key in month_columns.items():
            if not month_key or not _is_lesson_active_in_month(lesson_start, lesson_end, month_key):
                continue

            amount, charge_label = _month_charge_amount(row, column_index, regular_tuition)
            if amount <= 0:
                continue

            records.append(
                TuitionStudentMonth(
                    month=month_key,
                    student_name=student_name,
                    teacher_name=teacher_name,
                    payment_method=payment_method,
                    amount=amount,
                    regular_tuition=regular_tuition,
                    charge_label=charge_label,
                    lesson_start_date=lesson_start,
                    lesson_end_date=lesson_end,
                )
            )

    return records


def _payment_method_revenues(
    tuition_student_months: list[TuitionStudentMonth],
    session_collections: list[SessionCollection],
) -> list[PaymentMethodRevenue]:
    records: list[PaymentMethodRevenue] = []
    tuition_totals: dict[tuple[str, str], float] = defaultdict(float)
    for row in tuition_student_months:
        payment_method = row.payment_method or "미입력"
        tuition_totals[(row.month, payment_method)] += row.amount

    for (month_key, payment_method), amount in tuition_totals.items():
        records.append(
            PaymentMethodRevenue(
                month=month_key,
                payment_method=payment_method,
                amount=amount,
                source="tuition_monthly",
            )
        )

    session_totals: dict[tuple[str, str], float] = defaultdict(float)
    for row in session_collections:
        payment_method = _normalize_payment_method(row.payment_method)
        if not payment_method:
            continue
        session_totals[(row.month, payment_method)] += row.current_month_amount

    for (month_key, payment_method), amount in session_totals.items():
        records.append(
            PaymentMethodRevenue(
                month=month_key,
                payment_method=payment_method,
                amount=amount,
                source="session_collection",
            )
        )

    return records


def _clear_existing_data(db: Session) -> None:
    for model in (
        DataSync,
        TeacherSettlement,
        MonthlySettlement,
        TuitionRecord,
        TeacherProfile,
        SessionSettlement,
        SessionCollection,
        TuitionStudentMonth,
        PaymentMethodRevenue,
        ProductPrice,
        RateTableRow,
        TrialLesson,
    ):
        db.query(model).delete()
    db.commit()


def _upsert_student_records(db: Session, student_names: set[str]) -> None:
    existing = {
        row.student_name: row for row in db.query(StudentRecord).filter(StudentRecord.student_name.in_(student_names)).all()
    }
    now = datetime.utcnow()
    for student_name in sorted(student_names):
        if student_name in existing:
            existing[student_name].updated_at = now
            continue
        db.add(
            StudentRecord(
                student_name=student_name,
                parent_name=None,
                contact=None,
                status="수업중",
                notes=None,
                created_at=now,
                updated_at=now,
            )
        )


def _teacher_settlement_rows(worksheet) -> list[TeacherSettlement]:
    records: list[TeacherSettlement] = []
    current_month = ""

    for row in worksheet.iter_rows(values_only=True):
        marker = row[1] if len(row) > 1 else None
        marker_text = _text(marker)

        if "선생님 정산표" in marker_text:
            current_month = _month_from_value(marker_text)
            continue

        if not current_month or not marker_text:
            continue

        if marker_text in {"월별 정산", "회당 정산", "최종 합계", "선생님"}:
            continue

        if not isinstance(marker, str):
            continue

        records.append(
            TeacherSettlement(
                month=current_month,
                teacher_name=marker_text,
                student_count=_int(row[2] if len(row) > 2 else None),
                monthly_total_tuition=_float(row[3] if len(row) > 3 else None),
                monthly_trial_amount=_float(row[4] if len(row) > 4 else None),
                monthly_pretax_amount=_float(row[5] if len(row) > 5 else None),
                session_total_tuition=_float(row[6] if len(row) > 6 else None),
                session_trial_amount=_float(row[7] if len(row) > 7 else None),
                session_pretax_amount=_float(row[8] if len(row) > 8 else None),
                final_pretax_amount=_float(row[9] if len(row) > 9 else None),
                final_aftertax_amount=_float(row[10] if len(row) > 10 else None),
                settlement_date=_date(row[11] if len(row) > 11 else None),
            )
        )

    return records


def _monthly_settlement_rows(worksheet) -> list[MonthlySettlement]:
    records: list[MonthlySettlement] = []
    current_month = ""
    in_section = False

    for row in worksheet.iter_rows(values_only=True):
        marker = row[1] if len(row) > 1 else None
        marker_text = _text(marker)

        if isinstance(marker, (date, datetime)):
            current_month = _month_from_value(marker)
            in_section = False
            continue

        if marker_text == "선생님":
            in_section = True
            continue

        if not in_section or not current_month or not marker_text:
            continue

        if not isinstance(marker, str):
            continue

        records.append(
            MonthlySettlement(
                month=current_month,
                teacher_name=marker_text,
                student_count=_int(row[2] if len(row) > 2 else None),
                fee_rate=_float(row[3] if len(row) > 3 else None),
                first_payment=_float(row[4] if len(row) > 4 else None),
                recurring_payment=_float(row[5] if len(row) > 5 else None),
                special_payment=_float(row[6] if len(row) > 6 else None),
                refund_amount=_float(row[7] if len(row) > 7 else None),
                long_term_payment=_float(row[8] if len(row) > 8 else None),
                long_term_refund=_float(row[9] if len(row) > 9 else None),
                total_tuition=_float(row[10] if len(row) > 10 else None),
                trial_lesson_amount=_float(row[11] if len(row) > 11 else None),
                pretax_amount=_float(row[12] if len(row) > 12 else None),
            )
        )

    return records


def _tuition_records(worksheet) -> list[TuitionRecord]:
    records: list[TuitionRecord] = []

    for row in worksheet.iter_rows(min_row=5, values_only=True):
        teacher_name = _text(row[2] if len(row) > 2 else None)
        if not teacher_name or teacher_name == "홍길동":
            continue

        sequence_value = row[0] if len(row) > 0 else None
        if sequence_value == "ex)":
            continue

        records.append(
            TuitionRecord(
                sequence_no=_int(sequence_value) if sequence_value not in (None, "") else None,
                teacher_name=teacher_name,
                phone=_text(row[3] if len(row) > 3 else None) or None,
                email=_text(row[4] if len(row) > 4 else None) or None,
                birth_date_text=_text(row[5] if len(row) > 5 else None) or None,
                gender=_text(row[6] if len(row) > 6 else None) or None,
                education=_text(row[7] if len(row) > 7 else None) or None,
                major=_text(row[8] if len(row) > 8 else None) or None,
                teaching_experience=_text(row[9] if len(row) > 9 else None) or None,
                subject=_text(row[10] if len(row) > 10 else None) or None,
                available_grades=_text(row[11] if len(row) > 11 else None) or None,
            )
        )

    return records


def _teacher_profiles(records: Iterable[TuitionRecord], known_teacher_names: set[str]) -> list[TeacherProfile]:
    by_teacher: dict[str, dict[str, str | None]] = defaultdict(dict)

    for record in records:
        entry = by_teacher[record.teacher_name]
        for field in (
            "phone",
            "email",
            "birth_date_text",
            "gender",
            "education",
            "major",
            "teaching_experience",
            "subject",
            "available_grades",
        ):
            current_value = getattr(record, field)
            if current_value and not entry.get(field):
                entry[field] = current_value

    profiles = []
    for teacher_name in sorted(known_teacher_names):
        entry = by_teacher.get(teacher_name, {})
        profiles.append(
            TeacherProfile(
                teacher_name=teacher_name,
                phone=entry.get("phone"),
                email=entry.get("email"),
                birth_date_text=entry.get("birth_date_text"),
                gender=entry.get("gender"),
                education=entry.get("education"),
                major=entry.get("major"),
                teaching_experience=entry.get("teaching_experience"),
                subject=entry.get("subject"),
                available_grades=entry.get("available_grades"),
            )
        )
    return profiles


def _session_settlements(worksheet) -> list[SessionSettlement]:
    records: list[SessionSettlement] = []
    current_month = ""
    in_section = False

    for row in worksheet.iter_rows(values_only=True):
        marker = row[1] if len(row) > 1 else None
        marker_text = _text(marker)

        if isinstance(marker, (date, datetime)):
            current_month = _month_from_value(marker)
            in_section = False
            continue

        if marker_text == "선생님":
            in_section = True
            continue

        if not in_section or not current_month or not marker_text:
            continue

        if not isinstance(marker, str):
            continue

        records.append(
            SessionSettlement(
                month=current_month,
                teacher_name=marker_text,
                student_count=_int(row[2] if len(row) > 2 else None),
                first_payment_count=_int(row[3] if len(row) > 3 else None),
                first_payment_fee=_float(row[4] if len(row) > 4 else None),
                recurring_payment_count=_int(row[5] if len(row) > 5 else None),
                recurring_payment_fee=_float(row[6] if len(row) > 6 else None),
                refund_payment_count=_int(row[7] if len(row) > 7 else None),
                refund_payment_fee=_float(row[8] if len(row) > 8 else None),
                first_payment_commission=_float(row[9] if len(row) > 9 else None),
                recurring_payment_commission=_float(row[10] if len(row) > 10 else None),
                refund_payment_commission=_float(row[11] if len(row) > 11 else None),
            )
        )

    return records


def _header_month_column_map(header_row: tuple) -> dict[int, str]:
    month_columns: dict[int, str] = {}
    for index, value in enumerate(header_row):
        if isinstance(value, (date, datetime)):
            month_columns[index] = _month_from_value(value)
    return month_columns


def _append_trial_lesson(
    records: list[TrialLesson],
    seen: set[tuple[str, str, date]],
    *,
    teacher_name: str,
    student_name: str,
    trial_date: date,
    lesson_start: date | None,
    lesson_end: date | None,
    source: str,
) -> None:
    dedupe_key = (teacher_name, student_name, trial_date)
    if dedupe_key in seen:
        return
    seen.add(dedupe_key)
    records.append(
        TrialLesson(
            student_name=student_name,
            teacher_name=teacher_name,
            trial_lesson_date=trial_date,
            lesson_start_date=lesson_start,
            lesson_end_date=lesson_end,
            source=source,
        )
    )


def _trial_lessons_from_tuition_sheet(worksheet) -> list[TrialLesson]:
    """수업료 관리 AM열(시범수업 날짜) 및 월별 열에 들어 있는 날짜까지 수집."""
    records: list[TrialLesson] = []
    seen: set[tuple[str, str, date]] = set()
    header_row = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
    month_columns = _header_month_column_map(header_row)

    for row in worksheet.iter_rows(min_row=5, values_only=True):
        teacher_name = _text(row[TUITION_COL_TEACHER] if len(row) > TUITION_COL_TEACHER else None)
        student_name = _text(row[TUITION_COL_STUDENT] if len(row) > TUITION_COL_STUDENT else None)
        if not teacher_name or not student_name or student_name.startswith("▶"):
            continue

        lesson_start = _date(row[TUITION_COL_LESSON_START] if len(row) > TUITION_COL_LESSON_START else None)
        lesson_end = _date(row[TUITION_COL_LESSON_END] if len(row) > TUITION_COL_LESSON_END else None)

        trial_am = _date(row[TUITION_COL_TRIAL_DATE] if len(row) > TUITION_COL_TRIAL_DATE else None)
        if trial_am:
            _append_trial_lesson(
                records,
                seen,
                teacher_name=teacher_name,
                student_name=student_name,
                trial_date=trial_am,
                lesson_start=lesson_start,
                lesson_end=lesson_end,
                source="tuition_am",
            )

        for column_index in month_columns:
            cell = row[column_index] if len(row) > column_index else None
            if not isinstance(cell, (date, datetime)):
                continue
            trial_month_cell = _date(cell)
            if not trial_month_cell:
                continue
            _append_trial_lesson(
                records,
                seen,
                teacher_name=teacher_name,
                student_name=student_name,
                trial_date=trial_month_cell,
                lesson_start=lesson_start,
                lesson_end=lesson_end,
                source="tuition_month_cell",
            )

    return records


def _trial_lessons_from_session_sheet(worksheet) -> list[TrialLesson]:
    """회당 수금표 M열(시범수업 날짜) 및 월별 열의 날짜 값 수집."""
    records: list[TrialLesson] = []
    seen: set[tuple[str, str, date]] = set()
    header_row = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
    month_columns = _header_month_column_map(header_row)

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        teacher_name = _text(row[SESSION_COL_TEACHER] if len(row) > SESSION_COL_TEACHER else None)
        student_name = _text(row[SESSION_COL_STUDENT] if len(row) > SESSION_COL_STUDENT else None)
        if not teacher_name or not student_name:
            continue

        lesson_start = _date(row[SESSION_COL_LESSON_START] if len(row) > SESSION_COL_LESSON_START else None)
        lesson_end = _date(row[SESSION_COL_LESSON_END] if len(row) > SESSION_COL_LESSON_END else None)

        trial_m = _date(row[SESSION_COL_TRIAL_DATE] if len(row) > SESSION_COL_TRIAL_DATE else None)
        if trial_m:
            _append_trial_lesson(
                records,
                seen,
                teacher_name=teacher_name,
                student_name=student_name,
                trial_date=trial_m,
                lesson_start=lesson_start,
                lesson_end=lesson_end,
                source="session_m_column",
            )

        for column_index in month_columns:
            cell = row[column_index] if len(row) > column_index else None
            if not isinstance(cell, (date, datetime)):
                continue
            trial_month_cell = _date(cell)
            if not trial_month_cell:
                continue
            _append_trial_lesson(
                records,
                seen,
                teacher_name=teacher_name,
                student_name=student_name,
                trial_date=trial_month_cell,
                lesson_start=lesson_start,
                lesson_end=lesson_end,
                source="session_month_cell",
            )

    return records


def _merge_trial_lessons(
    tuition_trials: list[TrialLesson],
    session_trials: list[TrialLesson],
) -> list[TrialLesson]:
    merged: list[TrialLesson] = []
    seen: set[tuple[str, str, date]] = set()
    for row in [*tuition_trials, *session_trials]:
        key = (row.teacher_name, row.student_name, row.trial_lesson_date)
        if key in seen:
            continue
        seen.add(key)
        merged.append(row)
    return merged


def _student_names_from_tuition_sheet(worksheet) -> set[str]:
    names: set[str] = set()
    for row in worksheet.iter_rows(min_row=5, values_only=True):
        student_name = _text(row[15] if len(row) > 15 else None)
        if student_name and not student_name.startswith("▶"):
            names.add(student_name)
    return names


def _session_collections(worksheet) -> list[SessionCollection]:
    records: list[SessionCollection] = []
    header_row = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
    month_columns: dict[int, str] = {}
    for index, value in enumerate(header_row):
        if isinstance(value, (date, datetime)):
            month_columns[index] = _month_from_value(value)

    display_month = _month_from_value(worksheet["K2"].value)

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        teacher_name = _text(row[2] if len(row) > 2 else None)
        student_name = _text(row[3] if len(row) > 3 else None) or None
        if not teacher_name:
            continue

        display_amount = _float(row[11] if len(row) > 11 else None)
        display_sessions = _int(row[10] if len(row) > 10 else None)
        base_trial = _date(row[SESSION_COL_TRIAL_DATE] if len(row) > SESSION_COL_TRIAL_DATE else None)
        first_payment_amount = _float(row[15] if len(row) > 15 else None)
        lesson_start = _date(row[SESSION_COL_LESSON_START] if len(row) > SESSION_COL_LESSON_START else None)
        lesson_end = _date(row[SESSION_COL_LESSON_END] if len(row) > SESSION_COL_LESSON_END else None)

        shared_fields = {
            "row_number": _int(row[0] if len(row) > 0 else None) if row[0] not in (None, "") else None,
            "payment_method": _normalize_payment_method(row[1] if len(row) > 1 else None),
            "teacher_name": teacher_name,
            "student_name": student_name,
            "commission_rate": _float(row[4] if len(row) > 4 else None),
            "course": _text(row[5] if len(row) > 5 else None) or None,
            "weekly_frequency": _text(row[6] if len(row) > 6 else None) or None,
            "weekdays": _text(row[7] if len(row) > 7 else None) or None,
            "time_text": _text(row[8] if len(row) > 8 else None) or None,
            "product_name": _text(row[9] if len(row) > 9 else None) or None,
            "lesson_start_date": lesson_start,
            "lesson_end_date": lesson_end,
        }

        month_activity: dict[str, dict[str, Any]] = {}

        def ensure_month(month_key: str) -> dict[str, Any]:
            return month_activity.setdefault(
                month_key,
                {"amount": 0.0, "sessions": 0, "trial_date": None},
            )

        for column_index, month_key in month_columns.items():
            if display_month and month_key > display_month:
                continue

            cell = row[column_index] if len(row) > column_index else None
            if cell in (None, ""):
                continue

            entry = ensure_month(month_key)

            if isinstance(cell, (date, datetime)):
                entry["trial_date"] = _date(cell)
                continue

            label = _text(cell)
            if label in RECURRING_CHARGE_LABELS:
                if label == "첫달결제" and first_payment_amount > 0:
                    amount = first_payment_amount
                elif month_key == display_month and display_amount > 0:
                    amount = display_amount
                else:
                    amount = 0.0
                if amount > 0:
                    entry["amount"] = max(entry["amount"], amount)
                    entry["sessions"] = max(entry["sessions"], display_sessions)
            elif isinstance(cell, (int, float)) and cell > 0:
                entry["amount"] = max(entry["amount"], float(cell))
                entry["sessions"] = max(entry["sessions"], display_sessions)

        if base_trial:
            trial_month = _month_from_value(base_trial)
            if trial_month and (not display_month or trial_month <= display_month):
                entry = ensure_month(trial_month)
                if not entry["trial_date"]:
                    entry["trial_date"] = base_trial

        if not month_activity and display_month:
            month_activity[display_month] = {
                "amount": display_amount,
                "sessions": display_sessions,
                "trial_date": base_trial if base_trial and _month_from_value(base_trial) == display_month else None,
            }

        for month_key, entry in month_activity.items():
            if display_month and month_key > display_month:
                continue
            if not entry["trial_date"] and entry["amount"] <= 0 and entry["sessions"] <= 0:
                continue

            records.append(
                SessionCollection(
                    month=month_key,
                    current_month_amount=entry["amount"],
                    current_month_sessions=entry["sessions"],
                    trial_lesson_date=entry["trial_date"],
                    **shared_fields,
                )
            )

    return records


def _product_prices() -> list[ProductPrice]:
    records: list[ProductPrice] = []
    for table_name, rows in PRODUCT_PRICE_TABLES:
        for product_name, amount in rows:
            records.append(
                ProductPrice(
                    table_name=table_name,
                    product_name=product_name,
                    amount=float(amount),
                )
            )
    return records


def _rate_table_rows(worksheet) -> list[RateTableRow]:
    records: list[RateTableRow] = []
    current_section = ""
    row_order = 0

    for row in worksheet.iter_rows(values_only=True):
        values = list(row[:7])
        if not any(value not in (None, "") for value in values):
            continue

        if values[0] is None and isinstance(values[1], str) and not any(values[2:]):
            current_section = _text(values[1])
            continue

        row_type = "header" if values[0] is None else "data"
        records.append(
            RateTableRow(
                section_name=current_section,
                row_type=row_type,
                row_label=_text(values[0]),
                value_1=_text(values[1]) or None,
                value_2=_text(values[2]) or None,
                value_3=_text(values[3]) or None,
                value_4=_text(values[4]) or None,
                value_5=_text(values[5]) or None,
                value_6=_text(values[6]) or None,
                row_order=row_order,
            )
        )
        row_order += 1

    return records


def sync_workbook(
    db: Session,
    *,
    workbook_bytes: bytes | None = None,
    workbook_path: str | Path | None = None,
    source_name: str,
) -> None:
    if workbook_bytes is None and workbook_path is None:
        raise ValueError("workbook_bytes 또는 workbook_path 중 하나는 필요합니다.")

    if workbook_bytes is not None:
        workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
    else:
        workbook = load_workbook(filename=str(workbook_path), data_only=True)

    teacher_settlements = _teacher_settlement_rows(workbook["선생님 정산"])
    monthly_settlements = _monthly_settlement_rows(workbook["월별 정산"])
    tuition_records = _tuition_records(workbook["수업료 관리"])
    session_settlements = _session_settlements(workbook["회당 정산"])
    session_worksheet = workbook["회당 수금표"]
    session_collections = _session_collections(session_worksheet)
    tuition_worksheet = workbook["수업료 관리"]
    tuition_student_months = _tuition_student_months(tuition_worksheet)
    trial_lessons = _merge_trial_lessons(
        _trial_lessons_from_tuition_sheet(tuition_worksheet),
        _trial_lessons_from_session_sheet(session_worksheet),
    )
    payment_method_revenues = _payment_method_revenues(tuition_student_months, session_collections)
    product_prices = _product_prices()
    rate_table_rows = _rate_table_rows(workbook["단가표"])

    teacher_names = {
        record.teacher_name
        for record in teacher_settlements + monthly_settlements + session_settlements + session_collections
    }
    teacher_names.update(record.teacher_name for record in tuition_records)
    student_names = {record.student_name for record in session_collections if record.student_name}
    student_names.update(_student_names_from_tuition_sheet(tuition_worksheet))
    student_names.update(record.student_name for record in trial_lessons)
    student_names.update(record.student_name for record in tuition_student_months if record.student_name)
    teacher_profiles = _teacher_profiles(tuition_records, teacher_names)

    months = sorted({record.month for record in teacher_settlements if record.month})

    _clear_existing_data(db)
    db.add_all(teacher_settlements)
    db.add_all(monthly_settlements)
    db.add_all(tuition_records)
    db.add_all(teacher_profiles)
    db.add_all(session_settlements)
    db.add_all(session_collections)
    db.add_all(tuition_student_months)
    db.add_all(payment_method_revenues)
    db.add_all(product_prices)
    db.add_all(rate_table_rows)
    db.add_all(trial_lessons)
    _upsert_student_records(db, student_names)
    sync_teachers_and_enrollments(
        db,
        teacher_profiles=teacher_profiles,
        tuition_records=tuition_records,
        session_collections=session_collections,
    )
    sync_student_payment_methods(db, overwrite=True)
    db.add(
        DataSync(
            source_name=source_name,
            imported_at=datetime.utcnow(),
            notes=", ".join(months),
        )
    )
    db.commit()
